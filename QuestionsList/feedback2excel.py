'''
Created on Dec 21, 2023

@author: k2d2
'''
import sys
import os
from collections import OrderedDict
from shapes_and_collections.collections import col
sys.path.append(os.path.join(os.path.dirname(os.getcwd()), '360Survey/Db'))
print(sys.path)
from MySQLConnection import MySQLConnection
import getopt
import mysql.connector
from enum import Enum
from getpass import getpass
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.reader.excel import load_workbook
from dbinterface import *


db_host= 'localhost'
db_name= '360_survey_schema'
db_user = 'root'
db_password = 'MySql@123'

class ReportFormat:
  HdrFormat = Font(name='Calibri', size=10, bold=True, color='FF000000')
  HdrAlignment = Alignment(horizontal="center", vertical="center")
  SubHdrFormat = Font(name='Arial', size=9, bold=True, color='FF0000FF')
  SubHdrAlignment= Alignment(horizontal="center", vertical="center")

  ThinSide =  Side(border_style="thin", color="000000")
  ThickSide = Side(border_style = "thick", color="000000")
  MediumSide = Side(border_style = "medium", color="000000")
  
  HdrAllThickBorder = Border(top=ThickSide, left=ThickSide,
                             right=ThickSide, bottom=ThickSide)
  SubHdrAllBorder = Border(left=MediumSide, right=MediumSide, bottom=MediumSide)
  HdrTopThickBorder = Border(top=ThickSide, left=ThickSide, right=ThickSide)
  HdrBottomThickBorder = Border(left=ThickSide, right=ThickSide, bottom=ThickSide)
  HdrLeftAndBottomThickBorder = Border(left=ThickSide, bottom=ThickSide)
  HdrRightAndBottomThickBorder = Border(right=ThickSide, bottom=ThickSide)
  HdrLeftAndRightThickBorder = Border(left=ThickSide, right=ThickSide)
  HdrTopThickLeftThickBottomThinBorder = Border(top=ThickSide, left=ThickSide,
                                                bottom=ThinSide)
  HdrTopThickBottomThinBorder = Border(top=ThickSide, bottom=ThinSide)
  HdrTopThickRightThickBottomThinBorder = Border(top=ThickSide, left=ThickSide, 
                                                 right=ThickSide, bottom=ThinSide)
  HdrTopThickLeftThickRightThickBottomThinBorder = Border(top=ThickSide, left=ThickSide, 
                                                          right=ThickSide, bottom=ThinSide)
  HdrLeftThickRightThinBottomThickBorder = Border(left=ThickSide, right=ThinSide,
                                                bottom=ThickSide)
  HdrLeftThinRightThinBottomThickBorder = Border(left=ThinSide, right=ThinSide,
                                                bottom=ThickSide)
  HdrLeftThinRightThickBottomThickBorder = Border(left=ThinSide, right=ThickSide,
                                                bottom=ThickSide)    
  ThickBorder = Border(ThickSide)
  ThinBorder = Border(ThinSide)
  MediumBorder = Border(MediumSide)
  TopThickBorder = Border(top=ThickSide)
  RightThickBorder = Border(right=ThickSide)
  LeftThickBorder = Border(left=ThickSide)
  BottomThickBorder = Border(bottom=ThickSide)
  BottomMediumBorder = Border(bottom=MediumSide)
  BottomThinBorder = Border(bottom = ThinSide)
  TopThickRightThickBorder = Border(top=ThickSide, right=ThickSide)
  BottomThickRightThickBorder = Border(bottom=ThickSide, right=ThickSide)
  TopThickLeftThickBorder = Border(top=ThickSide, left=ThickSide)
  BottomThickLeftThickBorder = Border(bottom=ThickSide, left=ThickSide)
  AlignmentCenter= Alignment(horizontal="center", vertical="center")
  AlignmentRight = Alignment(horizontal="right")
  AlignmentLeft = Alignment(horizontal="left")
  HdrFill = PatternFill(start_color= '006B8E23', end_color='006B8E23', fill_type='solid')
  SubHdrFill = PatternFill(start_color= '00ADFF2F', end_color='00ADFF2F', fill_type='solid')
  HighlightFill = PatternFill(start_color= '00F0E68C', end_color='00F0E68C', fill_type='solid')


def WriteFeedbackReport(mysql_con_obj, wb, survey_id, eid):
  print("----------------------WriteFeedbackReport:[start]----------------------------------")
  feedback_map = OrderedDict()
  ws = wb.create_sheet("FeedbackReport", 0)
  print(f"Obtaining feedback by Surveyor type for employee={eid}, type={type(eid)} for survey={survey_id}\n")
  GetSurveyFeedbackForEmployeeBySurveyorType(mysql_con_obj, survey_id, eid, feedback_map)
  print(f"Obtaining QuestionListId of Survey={survey_id}")
  question_list_id = GetQuestionListIdOfSurvey(mysql_con_obj, survey_id)
  print(f"Obtaining Question List details for QuestionListId={question_list_id}")
  questionaire_map = OrderedDict()
  GetQuestionaireDetailsFromQuestionListId(mysql_con_obj, question_list_id, questionaire_map)  
  print(f"Creating Average Feedback received for each Evaluation Category by the employee {eid} in Survey {survey_id} from different Surveyor Types")
   
  row = 1
  cat_qid_col = 1
  cat_question_col = cat_qid_col + 1
  cat_qid_col_width = 8.0
  cat_question_col_width =  40.0
  feedback_col_width = 20.0
  cat_qid_col_code = get_column_letter(cat_qid_col)
  cat_question_col_code =  get_column_letter(cat_question_col)
  ws.column_dimensions[cat_qid_col_code].width = cat_qid_col_width
  ws.column_dimensions[cat_question_col_code].width = cat_question_col_width
  ws.merge_cells( cat_qid_col_code + str(row) + ':' + cat_question_col_code + str(row)) 
  ws.cell(row=row, column=cat_qid_col).value = "QuestionList:" + str(question_list_id)
  ws.cell(row=row, column=cat_qid_col).font = ReportFormat.HdrFormat
  ws.cell(row=row, column=cat_qid_col).alignment = ReportFormat.HdrAlignment
  ws.cell(row=row, column=cat_qid_col).border = ReportFormat.HdrAllThickBorder
  ws.cell(row=row, column=cat_question_col).border = ReportFormat.RightThickBorder
  ws.cell(row=row, column=cat_qid_col).fill = ReportFormat.HdrFill
  ws.cell(row=row + 1, column=cat_qid_col).value = "Number"
  ws.cell(row=row + 1, column=cat_qid_col).font = ReportFormat.SubHdrFormat
  ws.cell(row=row + 1, column=cat_qid_col).alignment = ReportFormat.SubHdrAlignment
  ws.cell(row=row + 1, column=cat_qid_col).border = ReportFormat.SubHdrAllBorder 
  ws.cell(row=row + 1, column=cat_qid_col).fill = ReportFormat.SubHdrFill
  ws.cell(row=row + 1, column=cat_question_col).value = "Question"
  ws.cell(row=row + 1, column=cat_question_col).font = ReportFormat.SubHdrFormat
  ws.cell(row=row + 1, column=cat_question_col).alignment = ReportFormat.SubHdrAlignment
  ws.cell(row=row + 1, column=cat_question_col).border = ReportFormat.SubHdrAllBorder
  ws.cell(row=row + 1, column=cat_question_col).fill = ReportFormat.SubHdrFill
  surveyor_count_total = 0
  feedback_start_col = 3
  col = feedback_start_col
  surveyor_type_row = 1
  surveyor_name_row = surveyor_type_row + 1
  eid2feedbackcol_map = {}
  qid2row_map = {}
  for surveyor_type_key in feedback_map:
    count = len(feedback_map[surveyor_type_key]['Surveyors'])
    surveyor_count_total += count
    print(f"Number of Surveyors of type:{surveyor_type_key} is={count}, total surveyor count={surveyor_count_total}")
    merge_start_col = col
    merge_start_col_code = get_column_letter(merge_start_col)
    for surveyor_eid in feedback_map[surveyor_type_key]['Surveyors']:
      print(f"SurveyorEid={surveyor_eid}, column={col}")
      col_code = get_column_letter(col)
      ws.column_dimensions[col_code].width = feedback_col_width
      ws.cell(row=surveyor_name_row, column=col).value = feedback_map[surveyor_type_key]['Surveyors'][surveyor_eid]['Name']
      ws.cell(row=surveyor_name_row, column=col).font = ReportFormat.SubHdrFormat
      ws.cell(row=surveyor_name_row, column=col).alignment = ReportFormat.SubHdrAlignment
      ws.cell(row=surveyor_name_row, column=col).border = ReportFormat.SubHdrAllBorder
      ws.cell(row=surveyor_name_row, column=col).fill = ReportFormat.SubHdrFill
      eid2feedbackcol_map[surveyor_eid] = col     
      col += 1
    ws.merge_cells(merge_start_col_code + str(surveyor_type_row) + ":" + get_column_letter(col - 1) + str(surveyor_type_row))
    ws.cell(row=surveyor_type_row, column=merge_start_col).value = surveyor_type_key
    ws.cell(row=surveyor_type_row, column=merge_start_col).font = ReportFormat.HdrFormat
    ws.cell(row=surveyor_type_row, column=merge_start_col).alignment = ReportFormat.HdrAlignment
    ws.cell(row=surveyor_type_row, column=merge_start_col).border = ReportFormat.HdrAllThickBorder
    ws.cell(row=surveyor_type_row, column=col - 1).border = ReportFormat.HdrAllThickBorder
    ws.cell(row=surveyor_type_row, column=merge_start_col).fill = ReportFormat.HdrFill
  r = row + 2
  small_font = Font(size=9)
  for category_id in questionaire_map['CategoryMap'].keys():
    print(f"CategoryId={category_id}, category={questionaire_map['CategoryMap'][category_id]['CategoryName']}")
    ws.merge_cells( cat_qid_col_code + str(r) + ':' + cat_question_col_code + str(r))
    ws.cell(row=r, column=cat_qid_col).value = questionaire_map['CategoryMap'][category_id]['CategoryName']
    ws.cell(row=r, column=cat_qid_col).font = ReportFormat.HdrFormat
    ws.cell(row=r, column=cat_qid_col).alignment = ReportFormat.HdrAlignment
    ws.cell(row=r, column=cat_qid_col).border = ReportFormat.HdrAllThickBorder
    ws.cell(row=r, column=cat_qid_col).border = ReportFormat.RightThickBorder
    ws.cell(row=r, column=cat_qid_col).fill = ReportFormat.HighlightFill
    r += 1
    qid2row_map[category_id] = {}
    for question_id in questionaire_map['CategoryMap'][category_id]['QuestionMap'].keys():
      print(f"\tQuestionId={question_id}, Question={questionaire_map['CategoryMap'][category_id]['QuestionMap'][question_id]}")
      ws.cell(row=r, column=cat_qid_col).value = question_id
      ws.cell(row=r, column=cat_question_col).value = questionaire_map['CategoryMap'][category_id]['QuestionMap'][question_id]
      ws.cell(row=r, column=cat_question_col).alignment = Alignment(wrap_text=True)
      ws.cell(row=r, column=cat_question_col).font =  small_font
      ws.row_dimensions[r].height = 25
      qid2row_map[category_id][question_id] = r
      r += 1
  
  for surveyor_type_key in feedback_map:
    print(f"Processing feedback from Surveyor Type={surveyor_type_key}")
    for surveyor_eid_key in feedback_map[surveyor_type_key]['Surveyors']:
      c = eid2feedbackcol_map[surveyor_eid_key]
      for (category_id, qid, rating) in feedback_map[surveyor_type_key]['Surveyors'][surveyor_eid_key]['Feedback']:
        r = qid2row_map[category_id][qid]
        print(f"Updating Feedback: row={r}, column={c}")
        print(f"SurveyorType={surveyor_type_key}, SurveyorEid={surveyor_eid_key} catid={category_id}, qid={qid}, rating={rating}")
        ws.cell(row=r, column=c).value = rating
  print("----------------------WriteFeedbackReport:[end]----------------------------------")

            
def Usage():
  print('Usage: feedbackreport.py [-h|--help] [-c|--conf <appconfig_file>] [-r|--report <excel_report_file>] [-s|--survey <survey_id>] <employee_id>\n')
  print('\tNote: Employee Id must be in quotes if the Id contains a space, e.g \'os 01\'')


if __name__ == '__main__':
  app_config_file = "excelreport.conf"
  report_file_name = None
  survey_id = 0
  eid = None
  db_host= 'localhost'
  db_name= '360_survey_schema'
  try:
    opts, args = getopt.getopt(sys.argv[1:],"hs:c:r:", ["help", "survey=", "conf=", "report="])
  except getopt.GetoptError:
    Usage()
    sys.exit(2)
  for opt, arg in opts:
    if opt in ('-h', '--help'):
      Usage()
      sys.exit(0)
    elif opt in ('-c', '--conf'):
      app_config_file = arg
      print("Config file specified:", app_config_file)
    elif opt in ('-s', '--survey'):
      survey_id = arg
    elif opt in ('-r', '--report'):
      report_file_name = arg
  if (len(args) < 1):
    print("Error: Employee Id argument is  mandatory")
    Usage()
    sys.exit(2)
  eid = str(args[0])
  if not survey_id:
    print(f"Error: Survey Id not provided, SurveyId is a mandatory input parameter\n")
    Usage()
    sys.exit(2)
  if not report_file_name:
    report_file_name = "eid_" + eid + "_survey_" + survey_id + ".xlsx"
    report_file_name = report_file_name.replace(" ", '_')
  print(f"Writing report to file=[{report_file_name}]")
  db_user=input("DB User:")
  db_password=getpass("DB Password:")
  wb = openpyxl.Workbook()
  mysql_con_obj = MySQLConnection(db_user, db_password, db_host, db_name)
  print("Successfully created mysql connection\n")
  WriteFeedbackReport(mysql_con_obj, wb, survey_id, eid)
  wb.save(report_file_name)
    