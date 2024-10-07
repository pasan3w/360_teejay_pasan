import sys
import getopt
import re
import os
from datetime import time
from collections import OrderedDict
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import configparser
from openpyxl.reader.excel import load_workbook

from getpass import getpass
import mysql.connector
from mysql.connector import errorcode
from _ast import arg


#DB connection details, should be provided in command line
DbUser = 'root'
DbPassword = '1qaz2wsx!@'
DbHost = 'localhost'
DbName = 'mysqldb'

class MySQLConnection():
  def __init__(self, user, password, host_name, db_name):
    self.User = user
    self.Password = password
    self.DbHost = host_name
    self.DbName = db_name
    self.connection = None
        
  def GetConnection(self):
    if (self.connection):
      return self.connection
    try:
      print("Opening connection to db:{} on Host:{}".format(self.DbHost, self.DbHost))
      self.connection = mysql.connector.connect(user=self.User, password=self.Password, 
                                                host=self.DbHost, database=self.DbName,
                                                auth_plugin='mysql_native_password')
    except mysql.connector.Error as err:
      self.connection = None
      if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
        print("Holy Shit, I am not allowed access, possible password mismatch, try again")
      elif err.errno == errorcode.ER_BAD_DB_ERROR:
        print("Not that bad error, trying to access Invalid database")
      else:
        print("Other error")
        print(err)

    return self.connection
    
    def __del__(self):
      print("Closing db connection")
      self.connection.close()
      self.connection = None
        
                
def ReadConfig(config_file):
  print("************ReadConfig:[start]************")
  config = configparser.ConfigParser()
  config.optionxform = str
  print("\tReading config file=", config_file)
  config.read(config_file)
  if 'company.employee.excelconfig' in config:
    print("\tCompany Employee List ExcelConfig: Number of keys loaded=", len(config['company.employee.excelconfig']))
    # for key in config['company.employee.excelconfig']:
    # print('\tKey=', key, ', Description=', config['company.employee.excelconfig'][key])
  if 'company.manager.excelconfig' in config:
    print("\tCompany Manager List ExcelConfig: Number of keys loaded=", len(config['company.manager.excelconfig']))
    # for key in config['company.manager.excelconfig']:
    # print('\tKey=', key, ', Description=', config['company.manager.excelconfig'][key])    
  print("************ReadConfig:[end]************")
  return config


def PrintConfig(app_config):
  print("************PrintConfig:[start]************")  
  sections = app_config.sections()
  for section in sections:
    print("--------Section=", section, "--------")
    options = app_config.options(section)
    for option in options:
      try:
        value = app_config.get(section, option) 
        print("\tAdd option={}, value={}".format(option, value))
      except:
        print("Error reading option=", option)
  print("************PrintConfig:[end]************")  


def BranchId2BranchName(db_con_obj, branch_id):
  print("************BranchId2BranchName[Start]************")
  print("BranchId=[{}]".format(branch_id))
  branch_name = ""
  con = db_con_obj.GetConnection()
  if (con.is_connected()):
    cursor = con.cursor()
    sql = "SELECT BranchName FROM Branch WHERE BranchID = %s;"
    cursor.execute(sql, (branch_id,))
    result = cursor.fetchone()
    if result:
      branch_name = result[0]
    else:
      print("BrancID:{} does not exist in Branch table".format(branch_id))
    cursor.close()
  print("************BranchId2BranchName[End]************")
  return branch_name

 
def DepartmentId2DepartmentName(db_con_obj, department_id):
  print("************DepartmentId2DepartmentName[Start]************")
  print("DepartmentId=[{}]".format(department_id))
  department_name = ""
  con = db_con_obj.GetConnection()
  if (con.is_connected()):
    cursor = con.cursor()
    sql = "SELECT DepartmentName FROM Department WHERE DepartmentID = %s;"
    cursor.execute(sql, (department_id,))
    result = cursor.fetchone()
    if result:
      department_name = result[0]
    else:
      print("DepartmentID:{} does not exist in Department table".format(department_id))
    cursor.close()
  print("************DepartmentId2DepartmentName[End]************")
  return department_name 


def JobTitleId2JobTitleName(db_con_obj, job_title_id):
  print("************JobTitleId2JobTitleName[Start]************")
  print("JobTitleId=[{}]".format(job_title_id))
  job_title_name = ""
  con = db_con_obj.GetConnection()
  if (con.is_connected()):
    cursor = con.cursor()
    sql = "SELECT JobTitleName FROM JobTitle WHERE JobTitleID = %s;"
    cursor.execute(sql, (job_title_id,))
    result = cursor.fetchone()
    if result:
      job_title_name = result[0]
    else:
      print("JobTitleID:{} does not exist in JobTitle table".format(job_title_id))
    cursor.close()
  print("************JobTitleId2JobTitleName[End]************")
  return job_title_name 


def GetSupervisors(db_con_obj, eid):
  print("************GetSupervisors:[start]************")
  print("Looking for supervisors of EID:[{}]".format(eid))
  print("Opening Db Connection")
  supervisors_str = ""
  con = db_con_obj.GetConnection()
  if not con:
    print("GetSupervisors: Failed to open DB Connection")
    sys.exit(2)
  
  if con.is_connected():
    cursor = con.cursor()
    sql = "SELECT SEID FROM Supervisor WHERE EID = %s"
    cursor.execute(sql, (eid,))
    result = cursor.fetchall()
    print("Result=", result)
    if result:
      supervisors_str = ','.join([str(i[0]) for i in result])
    cursor.close()
  print("GetSupervisors returning supervisor string:[{}]".format(supervisors_str) )
  print("************GetSupervisors:[end]************")
  return supervisors_str
    
    
def WriteEmployeeDataToExcel(db_con_obj, wb, app_config):
  print("************WriteEmployeeDataToExcel:[start]************")
  print("Opening Db Connection")
  con = db_con_obj.GetConnection()
  if not con:
    print("WriteEmployeeDataToExcel: Failed to open DB Connection")
    sys.exit(2)
  if con.is_connected():
    ws = wb.create_sheet(app_config['company.employee.excelconfig']['Sheet'], 0)
    ws.title = app_config['company.employee.excelconfig']['Sheet']
    employee_id_col = app_config['company.employee.excelconfig']['EmployeeIdCol']
    name_col = app_config['company.employee.excelconfig']['NameCol']
    email_address_col = app_config['company.employee.excelconfig']['EmailAddressCol']
    phone_number_col = app_config['company.employee.excelconfig']['PhoneNumberCol']
    branch_col = app_config['company.employee.excelconfig']['BranchCol']
    department_col = app_config['company.employee.excelconfig']['DepartmentCol']
    reporting_manager_col = app_config['company.employee.excelconfig']['ReportingManagerCol']
    job_title_col = app_config['company.employee.excelconfig']['JobTitleCol']
    ws.column_dimensions[employee_id_col].width = 16.0
    ws.column_dimensions[name_col].width = 40.0
    ws.column_dimensions[email_address_col].width = 40.0  
    ws.column_dimensions[phone_number_col].width = 20.0
    ws.column_dimensions[branch_col].width = 16.0
    ws.column_dimensions[department_col].width = 30.0
    ws.column_dimensions[reporting_manager_col].width = 30.0
    ws.column_dimensions[job_title_col].width = 50.0
    header_map = OrderedDict([(employee_id_col,'Employee Id') , (name_col,'Name'), (email_address_col , 'Email Address'),
                (phone_number_col,'Phone Number') , (branch_col,"Branch") , (department_col,'Department'),
                (reporting_manager_col,'Reporting Manager') , (job_title_col,'Job Title')])

    hdr_row_num = int(app_config['company.employee.excelconfig']['StartRow']) - 1
    ThickSide = Side(border_style = "thick", color="000000")
    for col_str in header_map.keys():
      hdr_col_num = column_index_from_string(col_str)                                 
      ws.cell(row=hdr_row_num, column=hdr_col_num).value = header_map[col_str]
      ws.cell(row=hdr_row_num, column=hdr_col_num).font = Font(name='Calibri', size=10, bold=True, color='FF000000')
      ws.cell(row=hdr_row_num, column=hdr_col_num).border = Border(top=ThickSide, left=ThickSide, right=ThickSide, bottom=ThickSide)
      ws.cell(row=hdr_row_num, column=hdr_col_num).alignment = Alignment(horizontal="center", vertical="center")
      ws.cell(row=hdr_row_num, column=hdr_col_num).fill = PatternFill(start_color= '0099CCFF', end_color='0099CCFF', fill_type='solid')
    employee_id_col_num = column_index_from_string(employee_id_col) 
    name_col_num = column_index_from_string(name_col) 
    email_address_col_num = column_index_from_string(email_address_col) 
    phone_number_col_num = column_index_from_string(phone_number_col) 
    branch_col_num = column_index_from_string(branch_col) 
    department_col_num = column_index_from_string(department_col) 
    reporting_manager_col_num = column_index_from_string(reporting_manager_col) 
    job_title_col_num = column_index_from_string(job_title_col) 
    row_num = hdr_row_num + 1       
    cursor = con.cursor()
    sql = """
      SELECT * FROM Employee;       
    """
    cursor.execute(sql)
    for row in cursor.fetchall():
      employee_id = row[0]
      name = row[1]
      branch_id = row[2]
      department_id = row[3]
      job_title_id = row[4]
      phone_number = str(row[5])
      email_address = row[6]
      print("From DB EmployeeId={}, Name={}, BranchId={}, DepartmentId={}, JobTitleId={}, PhoneNumber={}, Email={}".format(employee_id, name,
                                                                                                                          branch_id, department_id,
                                                                                                                          job_title_id, phone_number,
                                                                                                                          email_address))
      branch_name = BranchId2BranchName(db_con_obj, branch_id)
      department_name = DepartmentId2DepartmentName(db_con_obj, branch_id)
      job_title_name = JobTitleId2JobTitleName(db_con_obj, job_title_id)
      print("Writing 2 Excel: EmployeeId={}, Name={}, BranchName={}, DepartmentName={}, JobTitleName={}, PhoneNumber={}, Email={}".format(employee_id, name,
                                                                                                                                  branch_name, department_name,
                                                                                                                                  job_title_name, phone_number,
                                                                                                                                  email_address))
      print("Writing Excel row=", row_num)                                                                                                                                          
      ws.cell(row=row_num, column=employee_id_col_num).value = employee_id
      ws.cell(row=row_num, column=name_col_num).value = name
      ws.cell(row=row_num, column=branch_col_num).value = branch_name
      ws.cell(row=row_num, column=department_col_num).value = department_name
      ws.cell(row=row_num, column=reporting_manager_col_num).value = GetSupervisors(db_con_obj, employee_id)
      ws.cell(row=row_num, column=job_title_col_num).value = job_title_name
      ws.cell(row=row_num, column=email_address_col_num).value = email_address
      ws.cell(row=row_num, column=phone_number_col_num).value = phone_number
      row_num += 1
                 
  cursor.close()     
  print("************WriteEmployeeDataToExcel:[end]************")    

 
def WriteManagerDataToExcel(mysql_con_obj, wb, app_config):
  print("************WriteManagerDataToExcel:[start]************") 
  print("To be implemented")
  print("************WriteManagerDataToExcel:[end]************")   
  
def Usage():
  print('Usage: db2excel.py [--help] [-h|--host <host_name>] [-d|--db <db_name>] [-c|--conf <appconfig_file>] <employee_excel_file>')
    
                
if __name__ == "__main__":
  employee_list_file = None
  appconfigfile = 'load_employee_list.conf'
  db_host= 'localhost'
  db_name= '360_survey_schema'

  try:
    opts, args = getopt.getopt(sys.argv[1:],"h:d:c:", ["help", "host=", "db=","conf="])
  except getopt.GetoptError:
    Usage()
    sys.exit(2)
  print("args=", args)
  print("Number of args=", len(args))
  print("options=", opts)
  print("Number of opts=", len(opts))
  for opt, arg in opts:
    if opt in ('--help'):
      Usage()
      sys.exit(0)
    elif opt in ('-c', '--conf'):
      appconfigfile = arg
      print("Config file specified:", appconfigfile)
    elif opt in ('h', '--host'):
      db_host = arg
    elif opt in ('-d', '--db'):
      db_name = arg
  if (len(args) < 1):
    print("Error: Employee List Excel Filename argument is  mandatory")
    Usage()
    sys.exit(2)
  employee_list_file = args[0]
  db_user=input("DB User:")
  db_password=getpass("DB Password:")
  print("Db Host={}, Db Name={}, Db User={}, Db Password={}".format(db_host, db_name,
                                                                    db_user, db_password))
  
  print("Loading Application configuration from file: ", appconfigfile)
  AppConfig = ReadConfig(appconfigfile)
  PrintConfig(AppConfig)
  print("Loading employee list from file:", employee_list_file)
  employee_dic = OrderedDict()
  managers_dic = OrderedDict()
  wb = openpyxl.Workbook()
  mysql_con_obj = MySQLConnection(db_user, db_password, db_host, db_name)
  WriteEmployeeDataToExcel(mysql_con_obj, wb, AppConfig)
  WriteManagerDataToExcel(mysql_con_obj, wb, AppConfig)
  wb.save(employee_list_file)
