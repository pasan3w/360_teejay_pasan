'''
Created on Sep 25, 2023

@author: k2d2
'''

import sys
import openpyxl
from openpyxl.reader.excel import load_workbook


def LoadEmployeeData(excel_file_path, app_config, employee_dic):
  print("************LoadEmployeeList:[start]************")
  wb = load_workbook(filename = excel_file_path)
  if not wb:
    print("Error: Failed to open Excel file:", excel_file_path)
    sys.exit(2)
  ws = wb[app_config['company.employee.excelconfig']['Sheet']]
  if not ws:
    print("Error: Failed to open Employee data sheet:", app_config['company.employee.excelconfig']['sheet'])
    sys.exit(2)
    
  employee_id_col = app_config['company.employee.excelconfig']['EmployeeIdCol']  
  name_col = app_config['company.employee.excelconfig']['NameCol']
  email_address_col = app_config['company.employee.excelconfig']['EmailAddressCol']
  phone_number_col = app_config['company.employee.excelconfig']['PhoneNumberCol']
  branch_col = app_config['company.employee.excelconfig']['BranchCol']
  department_col = app_config['company.employee.excelconfig']['DepartmentCol']
  reporting_manager_col = app_config['company.employee.excelconfig']['ReportingManagerCol']
  job_title_col = app_config['company.employee.excelconfig']['JobTitleCol']
  print("Max row=", ws.max_row + 1)
  for row in range(int(app_config['company.employee.excelconfig']['StartRow']), ws.max_row + 1):
    srow = str(row)
    employee_id = str(ws[employee_id_col + srow].value).strip()
    if (employee_id is None):
      print("Empty employee id found, exiting")
      break
    name = ws[name_col + srow].value.strip() if ws[name_col + srow].value else None
    email_address = ws[email_address_col + srow].value.strip() if ws[email_address_col + srow].value else None
    phone_number = str(ws[phone_number_col + srow].value).strip() if ws[phone_number_col + srow].value else None
    branch = ws[branch_col + srow].value.strip() if ws[branch_col + srow].value else None
    department = ws[department_col + srow].value.strip() if ws[department_col + srow].value else None
    reporting_manager_eid = str(ws[reporting_manager_col + srow].value).strip() if ws[reporting_manager_col + srow].value else None
    job_title = ws[job_title_col + srow].value.strip() if ws[job_title_col + srow].value else None
    print("Row={}: EmployeeId={}, Name={}, EmailAddress={}, PhoneNumber={}, Branch={}, Department={}, Reporting ManagerEID={}, Job Title={}".format(row,
          employee_id, name, email_address, phone_number, branch, department, reporting_manager_eid, job_title))
    if (employee_id in employee_dic):
      print("Warning EmployeeId={} already exists in the Employee Dictionary".format(employee_id))
    employee_dic[employee_id]= {'Name' : name, 'EmailAddress' : email_address, 'PhoneNumber' : phone_number, 
                                'Branch' : branch, 'Department' : department, 'ReportingManagerEid' : reporting_manager_eid,
                                'JobTitle' : job_title}
  print("--------Printing Employee Dictionary--------")
  for key in employee_dic.keys():
    print("....EmployeeId={}....".format(key))
    print("\tName=", employee_dic[key]["Name"])
    print("\tEmailAddress=", employee_dic[key]["EmailAddress"])
    print("\tPhoneNumber=", employee_dic[key]["PhoneNumber"])
    print("\tBranch=", employee_dic[key]["Branch"])
    print("\tDepartment=", employee_dic[key]["Department"])
    print("\tReportingManagerEid=", employee_dic[key]["ReportingManagerEid"])    
   
  wb.close()
  print("************LoadEmployeeList:[end]************")
  

def LoadManagerData(excel_file_path, app_config, employee_dic, manager_dic):
  print("************LoadManagerData:[start]************")
  wb = load_workbook(filename = excel_file_path)
  if not wb:
    print("Error: Failed to open Excel file:", excel_file_path)
    sys.exit(2)
  ws = wb[app_config['company.manager.excelconfig']['Sheet']]
  if not ws:
    print("Error: Failed to open Manager data sheet:", app_config['company.manager.excelconfig']['sheet'])
    sys.exit(2)
  job_role_col = app_config['company.manager.excelconfig']['JobRoleCol']
  employee_id_col = app_config['company.manager.excelconfig']['EmployeeIdCol']  
  name_col = app_config['company.manager.excelconfig']['CandidateCol']
  reporting_managers_col = app_config['company.manager.excelconfig']['ReportingManagersCol']
  direct_reports_col = app_config['company.manager.excelconfig']['DirectReportsCol']
  peer_list_col = app_config['company.manager.excelconfig']['PeerListCol']
  print("Max row=", ws.max_row + 1)
  for row in range(int(app_config['company.manager.excelconfig']['StartRow']), ws.max_row + 1):
    srow = str(row)
    employee_id = str(ws[employee_id_col + srow].value).strip()
    if not employee_id:
      print("Empty employee id not found, exiting")
      break
    else:
      print("EmployeeId=[{}]".format(employee_id))
    name = ws[name_col + srow].value.strip() if ws[name_col + srow].value else None
    job_role = ws[job_role_col + srow].value.strip() if ws[job_role_col + srow].value else None
    reporting_managers_str = str(ws[reporting_managers_col + srow].value).strip() if ws[reporting_managers_col + srow].value else None
    direct_reports_str = str(ws[direct_reports_col + srow].value).strip() if ws[direct_reports_col + srow].value else None
    peer_list_str = str(ws[peer_list_col + srow].value).strip() if ws[peer_list_col + srow].value else None
    print("Row={}: EmployeeId={}, Candidate={}, JobRole={}, ReportingManagers={}, DirectReports=[{}], PeerList=[{}]".format(row,
          employee_id, name, job_role, reporting_managers_str, direct_reports_str, peer_list_str))
    if (employee_id not in employee_dic):
      print("Error: EmployeeId={} not present in the Employee Dictionary", employee_id)
      sys.exit(2)
    if employee_id not in manager_dic:
      manager_dic[employee_id] = {'Name' : name, 'IntegratedRMs' : [], 'IntegratedDRs' : [],
                                  'IntegratedPeerList' : [], 
                                  'job_roles' : { job_role : {'ReportingManagers' : reporting_managers_str,
                                                              'DirectReports' : direct_reports_str, 'PeerList' : peer_list_str}}}
    else:
      manager_dic[employee_id]['job_roles'][job_role] = {'ReportingManagers' : reporting_managers_str,
                      'DirectReports' : direct_reports_str, 'PeerList' : peer_list_str}
  print("--------Printing Managers Dictionary--------")
  for key in manager_dic.keys():
    print("----EmployeeId={}----".format(key))
    print("Name=", employee_dic[key]["Name"])
    job_roles = manager_dic[key]["job_roles"].keys()
    for job_role in job_roles:
      print("JobRole=", job_role)
      print("\tReportingManagers=", manager_dic[key]["job_roles"][job_role]["ReportingManagers"])
      if manager_dic[key]["job_roles"][job_role]["ReportingManagers"]:
        for rm_eid in str(manager_dic[key]["job_roles"][job_role]["ReportingManagers"]).split(','):
          rm_eid = rm_eid.strip()
          if rm_eid not in manager_dic[key]['IntegratedRMs']:
            manager_dic[key]['IntegratedRMs'].append(rm_eid)
      print("\tDirectReports=", manager_dic[key]["job_roles"][job_role]["DirectReports"])
      if manager_dic[key]["job_roles"][job_role]["ReportingManagers"]:
        for dr_eid in str(manager_dic[key]["job_roles"][job_role]["DirectReports"]).split(','):
          dr_eid = dr_eid.strip()
          if dr_eid not in manager_dic[key]['IntegratedDRs']:
            manager_dic[key]['IntegratedDRs']     
      print("\tPeerList=", manager_dic[key]["job_roles"][job_role]["PeerList"])
      if manager_dic[key]["job_roles"][job_role]["PeerList"]:
        for pr_eid in str(manager_dic[key]["job_roles"][job_role]["PeerList"]).split(','):
          pr_eid = pr_eid.strip()
          if pr_eid not in manager_dic[key]['IntegratedPeerList']:
            manager_dic[key]['IntegratedPeerList']

  wb.close()
  print("************LoadManagerData:[end]************")

