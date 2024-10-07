'''
Created on Sep 29, 2023

@author: k2d2
'''


import sys
import getopt
import os
import shutil
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from collections import OrderedDict
import configparser


def ReadConfig(config_file):
  print("************ReadConfig:[start]************")
  config = configparser.ConfigParser()
  config.optionxform = str
  print("\tReading config file=", config_file)
  config.read(config_file)
  if 'company.employee.excelconfig' in config:
    print("\tCompany Employee List ExcelConfig: Number of keys loaded=", len(config['company.employee.excelconfig']))
    # for key in config['company.employee.excelconfig']:
    # print('\tKey=', key, ', Description=', config['company.employee.excelconfig'][key])
  if 'company.manager.excelconfig' in config:
    print("\tCompany Manager List ExcelConfig: Number of keys loaded=", len(config['company.manager.excelconfig']))
    # for key in config['company.manager.excelconfig']:
    # print('\tKey=', key, ', Description=', config['company.manager.excelconfig'][key])
  if 'company.excelconfig' in config:
    print("\tExcelconfig: Number of keys loaded=", len(config['company.excelconfig']))
        
  print("************ReadConfig:[end]************")
  return config


def TransformValues(wb, app_config):
  print("--------------------TransformValues:[start]-------------------------------")
  edic = dict.fromkeys(app_config['company.excelconfig']['EliminateValues'].split(','), '')
  print("Number of keys in the Transform dic=", len(edic))
  print("Keys=[{}]".format(edic.keys()))
  print("----Transforming Employee values-------")
  ws = wb[app_config['company.employee.excelconfig']['Sheet']]
  print("Maxrows={}, Maxcol={}".format(ws.max_row, ws.max_column))
  for row in ws.iter_rows(min_row = int(app_config['company.employee.excelconfig']['StartRow']), max_row = ws.max_row, max_col = ws.max_column):
    for cell in row:
      print("Cell={}, Value={}".format(cell, cell.value))
      if cell.value in edic:
        print("Transforming value of cell:{} Value={}".format(cell, cell.value))
        cell.value = ""
  print("----Transforming Manager values-------")
  ws = wb[app_config['company.manager.excelconfig']['Sheet']]
  print("Maxrows={}, Maxcol={}".format(ws.max_row, ws.max_column))
  for row in ws.iter_rows(min_row = int(app_config['company.employee.excelconfig']['StartRow']), max_row = ws.max_row, max_col = ws.max_column):
    for cell in row:
      print(cell.value)
      if cell.value in edic:
        print("Transforming value of cell:{} Value={}".format(cell, cell.value))
        cell.value = ""
  print("--------------------TransformValues:[end]-------------------------------")
                            

def FixEmployeeDataSheet(wb, app_config, employee_dic):
  print("************FixEmployeeDataSheet:[start]************")
  ws = wb[app_config['company.employee.excelconfig']['Sheet']]
  if not ws:
    print("Error: Failed to open Employee data sheet:", app_config['company.employee.excelconfig']['Sheet'])
    sys.exit(2)    
  employee_id_col = app_config['company.employee.excelconfig']['EmployeeIdCol']  
  name_col = app_config['company.employee.excelconfig']['NameCol']
  branch_col = app_config['company.employee.excelconfig']['BranchCol']
  department_col = app_config['company.employee.excelconfig']['DepartmentCol']
  reporting_manager_col = app_config['company.employee.excelconfig']['ReportingManagerCol']
  job_title_col = app_config['company.employee.excelconfig']['JobTitleCol']
  email_address_col = app_config['company.employee.excelconfig']['EmailAddressCol']
  phone_number_col = app_config['company.employee.excelconfig']['PhoneNumberCol']
  
  branch_col_num = column_index_from_string(branch_col)
  print("Inserting Branch column at Col={}, index={}".format(branch_col, branch_col_num))
  ws.insert_cols(branch_col_num)
  ws.column_dimensions[branch_col].width = 16.0
  ws.column_dimensions[department_col].width = 26.0
  ws.column_dimensions[reporting_manager_col].width = 26.0
  ws.column_dimensions[job_title_col].width = 46.0
  ws.column_dimensions[email_address_col].width = 36.0
  ws.column_dimensions[phone_number_col].width = 26.0
  
  header_row_num = int(app_config['company.employee.excelconfig']['StartRow']) - 1
  ws.cell(row=header_row_num, column=branch_col_num).value = 'Branch'
  ws.cell(row=header_row_num, column=branch_col_num).font = Font(name='Times New Roman', size=12, color='FF000000')
  ThickSide = Side(border_style = "thick", color="000000")
  ThinSide = Side(border_style = "thin", color="000000")
  #ws.cell(row=header_row_num, column=branch_col_num).border = Border(top=ThickSide, left=ThickSide, right=ThickSide, bottom=ThickSide)
  ws.cell(row=header_row_num, column=branch_col_num).border = Border(top=ThinSide, left=ThinSide, right=ThinSide, bottom=ThinSide)
  ws.cell(row=header_row_num, column=branch_col_num).fill = PatternFill(start_color= '00b4c7e7', end_color='00b4c7e7', fill_type='solid')
  for row in range(int(app_config['company.employee.excelconfig']['StartRow']), ws.max_row + 1):
    srow = str(row)
    employee_id = str(ws[employee_id_col + srow].value).strip()
    if (employee_id is None):
      print("Empty employee id found, exiting")
      break
    print("Adding EmployeeId=[{}] to Dictionary".format(employee_id))
    employee_dic[employee_id] = {}
    employee_dic[employee_id]['Name'] = ws[name_col + srow].value.strip()
    employee_dic[employee_id]['ReportingManagers'] = []
    division_and_branch_ary = ws[department_col + srow].value.split('-')
    print("Branch=[{}], Division=[{}]".format(division_and_branch_ary[1].strip(), division_and_branch_ary[0].strip()))
    ws[branch_col + srow] = division_and_branch_ary[1].strip()
    ws[department_col + srow] = division_and_branch_ary[0].strip()
    reporting_manager = str(ws[reporting_manager_col + srow].value).strip()
    if reporting_manager:
      employee_dic[employee_id]['ReportingManagers'].append(reporting_manager)
  print("Number of Employees in Employee Dictionary=", len(employee_dic))
  print("************FixEmployeeDataSheet:[end]************")


def UpdateEmployeeDataSheetReportingManagers(wb, app_config, employee_dic):
  print("************UpdateEmployeeDataSheetReportingManagers:[start]************")
  ws = wb[app_config['company.employee.excelconfig']['Sheet']]
  if not ws:
    print("Error: Failed to open Employee data sheet:", app_config['company.employee.excelconfig']['Sheet'])
    sys.exit(2)    
  employee_id_col = app_config['company.employee.excelconfig']['EmployeeIdCol']  
  reporting_manager_col = app_config['company.employee.excelconfig']['ReportingManagerCol']
  for row in range(int(app_config['company.employee.excelconfig']['StartRow']), ws.max_row + 1):
    srow = str(row)
    employee_id = str(ws[employee_id_col + srow].value).strip()
    print("updating row={}, column={} employee_id={}".format(row, reporting_manager_col, employee_id))
    if not employee_id:
      print("Empty employee_id found, breaking")
      break
    ws[reporting_manager_col + srow].value = ','.join(employee_dic[str(employee_id)]['ReportingManagers'])
  print("************UpdateEmployeeDataSheetReportingManagers:[end]************")


def FixManagerDataSheet(wb, app_config, employee_dic):
  print("************FixManagerDataSheet:[start]************")
  ws = wb[app_config['company.manager.excelconfig']['Sheet']]
  if not ws:
    print("Error: Failed to open Manager data sheet:", app_config['company.manager.excelconfig']['Sheet'])
    sys.exit(2)    

  employee_id_col = app_config['company.manager.excelconfig']['EmployeeIdCol']
  reporting_managers_col = app_config['company.manager.excelconfig']['ReportingManagersCol']  
  direct_reports_col = app_config['company.manager.excelconfig']['DirectReportsCol']
  candidate_col = app_config['company.manager.excelconfig']['CandidateCol']
  peer_list_col = app_config['company.manager.excelconfig']['PeerListCol'] 
  reporting_managers_col_num = column_index_from_string(reporting_managers_col)
  print("Inserting Reporting Managers column at Col={}, index={}".format(reporting_managers_col, reporting_managers_col_num))
  ws.insert_cols(reporting_managers_col_num)
  ws.column_dimensions[candidate_col].width = 30.0
  ws.column_dimensions[reporting_managers_col].width = 26.0
  ws.column_dimensions[peer_list_col].width = 30.0
  header_row_num = int(app_config['company.employee.excelconfig']['StartRow']) - 1
  reporting_managers_col_num = column_index_from_string(reporting_managers_col)
  ws.cell(row=header_row_num, column=reporting_managers_col_num).value = 'Reporting Managers' 
  ws.cell(row=header_row_num, column=reporting_managers_col_num).font = Font(name='Times New Roman', size=14, color='FF000000')
  ThickSide = Side(border_style = "thick", color="000000")
  ws.cell(row=header_row_num, column=reporting_managers_col_num).border = Border(top=ThickSide, left=ThickSide, right=ThickSide, bottom=ThickSide)
  ws.cell(row=header_row_num, column=reporting_managers_col_num).fill = PatternFill(start_color= '00b4c7e7', end_color='00b4c7e7', fill_type='solid')
  for row in range(int(app_config['company.employee.excelconfig']['StartRow']), ws.max_row + 1):
    srow = str(row)
    if not ws[employee_id_col + srow].value:
      print("Empty employee id found, exiting")
      break
    employee_id = str(ws[employee_id_col + srow].value).strip()
    direct_reports_str = str(ws[direct_reports_col + srow].value)
    if direct_reports_str:
      direct_reports_ary = direct_reports_str.strip().split(',')
      for dr_id in direct_reports_ary:
        st_dr_id = dr_id.strip()
        if st_dr_id in employee_dic:
          print("AddRM: Employee=[{}] RM=[{}]".format(st_dr_id, employee_id))
          employee_dic[st_dr_id]['ReportingManagers'].append(employee_id)
        else:
          print("Strange: Employee=[{}] RM=[{}], The DR name does not exist, Should not happen".format(st_dr_id, employee_id))
  print("************FixManagerDataSheet:[end]************")


def UpdateManagerSheetReportingManagers(wb, app_config, employee_dic):
  print("************UpdateManagerSheetReportingManagers:[start]************")
  ws = wb[app_config['company.manager.excelconfig']['Sheet']]
  if not ws:
    print("Error: Failed to open Manager data sheet:", app_config['company.manager.excelconfig']['Sheet'])
    sys.exit(2)    
  employee_id_col = app_config['company.manager.excelconfig']['EmployeeIdCol']  
  reporting_manager_col = app_config['company.manager.excelconfig']['ReportingManagersCol']
  for row in range(int(app_config['company.manager.excelconfig']['StartRow']), ws.max_row + 1):
    srow = str(row)
    if not ws[employee_id_col + srow].value:
      print("Empty employee_id found, breaking")
      break      
    employee_id = str(ws[employee_id_col + srow].value).strip()
    print("updating row={}, column={} employee_id={}".format(row, reporting_manager_col, employee_id))
    ws[reporting_manager_col + srow].value = ','.join(employee_dic[employee_id]['ReportingManagers'])
  print("************UpdateManagerSheetReportingManagers:[end]************")

  
def EliminateDuplicateReportingManagers(employee_dic):
  print("----------------------EliminateDuplicateReportingManagers:[start]---------------------")
  print("Number of employees in the employee list=", len(employee_dic))
  for employee_id in employee_dic:
    print("Pre-duplicate elimination RMs=[", employee_dic[employee_id]['ReportingManagers'], "]")
    uniq_rm_list = list(OrderedDict.fromkeys(employee_dic[employee_id]['ReportingManagers']))
    employee_dic[employee_id]['ReportingManagers'] = uniq_rm_list
    print("After-duplicate elimination RMs=[", employee_dic[employee_id]['ReportingManagers'], "]")

     
def BackupExcelFile(file_path):
  path_parts = os.path.split(file_path)
  file_parts = path_parts[1].split('.')
  backup_file = os.path.join(path_parts[0], file_parts[0] + "_save." + file_parts[1])
  print("Copying file:{} to:{}".format(file_path, backup_file))
  shutil.copy(file_path, backup_file)
  
    
def Usage():
  print('Usage: fixdata.py [-h|--help] [-c|--conf <appconfig_file>] <employee_excel_file>')
    
                
if __name__ == "__main__":
  employee_list_file = None
  appconfigfile = 'fixdata.conf'

  try:
    opts, args = getopt.getopt(sys.argv[1:],"hc:", ["help","conf="])
  except getopt.GetoptError:
    Usage()
    sys.exit(2)
  print("args=", args)
  print("Number of args=", len(args))
  print("options=", opts)
  print("Number of opts=", len(opts))
  for opt, arg in opts:
    if opt in ('-h', '--help'):
      Usage()
      sys.exit(0)
    elif opt in ('-c', '--conf'):
      appconfigfile = arg
      print("Config file specified:", appconfigfile)
  if (len(args) < 1):
    print("Error: Employee List Excel Filename argument is  mandatory")
    Usage()
    sys.exit(2)
  employee_list_file = args[0]
  app_config = ReadConfig(appconfigfile)
  BackupExcelFile(employee_list_file)
  wb = load_workbook(filename = employee_list_file)
  if not wb:
    print("Error: Failed to open Excel file:", employee_list_file)
    sys.exit(2)
  TransformValues(wb, app_config)
  employee_dic = OrderedDict()
  managers_dic = OrderedDict()
  FixEmployeeDataSheet(wb, app_config, employee_dic)
  FixManagerDataSheet(wb, app_config, employee_dic)
  EliminateDuplicateReportingManagers(employee_dic)
  UpdateEmployeeDataSheetReportingManagers(wb, app_config, employee_dic)
  UpdateManagerSheetReportingManagers(wb, app_config, employee_dic) 
  wb.save(employee_list_file)
